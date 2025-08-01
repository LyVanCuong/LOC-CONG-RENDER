
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import os

def filter_excel_data(file_sv, file_cong, output_folder):
    df_sv = pd.read_excel(file_sv, sheet_name=0)
    df_cong = pd.read_excel(file_cong, sheet_name="Sheet1")

    if '工号' in df_sv.columns:
        df_sv.rename(columns={'工号': 'MaNhanVien'}, inplace=True)
    if 'Tên Trường' in df_sv.columns:
        df_sv.rename(columns={'Tên Trường': 'TenTruong'}, inplace=True)
    if '立讯工号' in df_cong.columns:
        df_cong.rename(columns={'立讯工号': 'MaNhanVien'}, inplace=True)

    df_sv['MaNhanVien'] = df_sv['MaNhanVien'].astype(str).str.strip()
    df_cong['MaNhanVien'] = df_cong['MaNhanVien'].astype(str).str.strip()

    df_merged = pd.merge(df_cong, df_sv[['MaNhanVien', 'TenTruong']], on='MaNhanVien', how='left')
    danh_sach_truong = df_merged['TenTruong'].dropna().unique()

    output_files = []

    for ten_truong in danh_sach_truong:
        df_loc = df_merged[df_merged['TenTruong'] == ten_truong].copy()
        df_loc = df_loc.sort_values(by='MaNhanVien', ascending=True)
        df_loc.insert(0, 'STT', range(1, len(df_loc) + 1))

        if 'TenTruong' in df_loc.columns:
            df_loc.drop(columns=['TenTruong'], inplace=True)

        safe_name = "_".join(ten_truong.strip().split())
        output_file = os.path.join(output_folder, f"ket_qua_loc_{safe_name}.xlsx")
        df_loc.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        font_default = Font(name='Times New Roman', size=10)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        header = [cell.value for cell in ws[1]]
        try:
            col_gio_len = next(i for i, h in enumerate(header) if h and ("上班卡" in str(h) or "QUÉT LÊN CA" in str(h))) + 1
            col_gio_xuong = next(i for i, h in enumerate(header) if h and ("下班卡" in str(h) or "QUÉT XUỐNG CA" in str(h))) + 1
        except StopIteration:
            wb.save(output_file)
            output_files.append(output_file)
            continue

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            len_value = row[col_gio_len - 1].value
            xuong_value = row[col_gio_xuong - 1].value
            if (len_value in [None, ""]) and (xuong_value in [None, ""]):
                for cell in row:
                    cell.fill = yellow_fill
            for cell in row:
                cell.font = font_default
                cell.border = border_thin
                cell.alignment = center_align

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in [8, 9]:
                if col_idx <= len(row):
                    row[col_idx - 1].number_format = 'dd/mm/yyyy'
            if len(row) >= 9:
                row[8].fill = yellow_fill

        for cell in ws[1]:
            cell.font = font_default
            cell.border = border_thin
            cell.alignment = center_align

        ws.delete_cols(2)
        ws.delete_rows(1)

        header_row = [
    "序号\nSTT", "人员编号\nMÃ THẺ", "立讯工号\nMÃ NHÂN VIÊN", "姓名\nHỌ VÀ TÊN",
    "组织单位\nBỘ PHẬN", "部门ID\nMÃ BỘ PHẬN", "入职日期\nNGÀY VÀO LÀM",
    "考勤日期\nNGÀY XUẤT CÔNG", "班次\nCA LÀM", "计划上班时间\nTHỜI GIAN CA LÀM",
    "上班卡\nQUẸT LÊN CA", "下班卡\nQUẸT XUỐNG CA"
]

        ws.insert_rows(1)
        for idx, value in enumerate(header_row, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = value
            cell.font = font_default
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color="AFAFAF", end_color="AFAFAF", fill_type="solid")

        wb.save(output_file)
        output_files.append(output_file)

    return output_files
